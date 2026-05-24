"""
Excel (.xlsx) ingestion for the clinical-case batch dataset.

Maps ONE spreadsheet row -> ONE clinical case. Each case's ``input_text`` is built by
**labeled-section concatenation**: every non-empty content column becomes a labeled
section (``"<Column name>:\n<value>"``), in sheet column order. This is deliberately NOT
regex header splitting (that is for free-text uploads) — a spreadsheet already has explicit
structure, and feeding the column labels to the pipeline gives the extraction layers
unambiguous anchors.

Design notes:
- ``openpyxl`` in ``read_only`` mode (streaming iterator) so a 156x23 sheet with large
  cells never materializes fully in memory (RISK_REGISTER N-4).
- The key column is ``ID`` (e.g. ``OLC-CHI-000123456``); it becomes ``case_label`` and is
  NOT emitted as a content section.
- Always-empty columns are dropped two ways: (1) an explicit, auditable drop-set of the
  known-empty headers, and (2) per-row empty-cell skipping — so even an unexpected stray
  empty column never produces a dangling ``Label:`` with no body.
- Cell text is cleaned: ``_x000D_`` escapes stripped, CR normalized, tabs->spaces, bullets
  normalized, meaningful embedded newlines preserved.

No FastAPI/DB imports here on purpose — keep this layer pure and unit-testable.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

# --- Column policy ---------------------------------------------------------------------

#: Header (normalized) used as the stable per-case key / label.
KEY_COLUMN = "ID"

#: Expected format of the ID column values (Swiss clinical export). Non-matching IDs are
#: still ingested but flagged in the validation report.
ID_PATTERN = re.compile(r"^OLC-CHI-\d{9}$")

#: Columns that are always-empty in the dataset and carry no clinical text. Dropped
#: explicitly for determinism; per-row empty-skipping (below) is the robust fallback.
#: Matching is on the normalized (stripped) header, case-insensitive.
DROP_COLUMNS = frozenset(
    h.casefold()
    for h in {
        "SOFA Score 1",
        "SOFA Score 2",
        "SOFA Score 3",
        "SOFA Score 4",
        "kardiovaskuläre Risikofaktoren",
        "Fahreignung / Fahrfähigkeit",
        "Labor",
    }
)

#: Any column whose normalized header starts with one of these is also dropped (guards
#: against SOFA columns named "SOFA Score" / "SOFA Score 1.0" etc.).
DROP_PREFIXES = ("sofa score",)

#: Columns considered "core" clinical text. A row with ALL of these empty is flagged as a
#: likely-empty case in the validation report (still ingested unless completely blank).
CORE_COLUMNS = frozenset(
    h.casefold()
    for h in {
        "Diagnosen",
        "Operationen",
        "Beurteilung / Verlauf",
        "Procedere",
    }
)

#: Preferred sheet name; falls back to auto-detection by header if absent.
PREFERRED_SHEET = "Export Berichte"


class ExcelParseError(ValueError):
    """Raised when the workbook cannot be parsed (corrupt, no usable sheet/ID column)."""


@dataclass
class CaseRecord:
    """One spreadsheet row mapped to one clinical case."""

    case_label: str  # the ID (or a "Row N" fallback)
    input_text: str  # labeled-section concatenation
    row_number: int  # 1-based data row index (diagnostics only)
    raw_id: str | None = None
    warnings: list[str] = field(default_factory=list)


@dataclass
class ValidationReport:
    """Summary surfaced to the uploader before the job is enqueued."""

    total_data_rows: int = 0
    valid_cases: int = 0
    skipped_blank_rows: int = 0
    duplicate_ids: list[str] = field(default_factory=list)
    missing_ids: list[int] = field(default_factory=list)  # row numbers with no ID
    invalid_id_format: list[str] = field(default_factory=list)
    empty_core_rows: list[str] = field(default_factory=list)  # labels lacking core text
    sheet_name: str = ""
    columns_used: list[str] = field(default_factory=list)
    columns_dropped: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    def as_dict(self) -> dict:
        return {
            "total_data_rows": self.total_data_rows,
            "valid_cases": self.valid_cases,
            "skipped_blank_rows": self.skipped_blank_rows,
            "duplicate_ids": self.duplicate_ids,
            "missing_ids": self.missing_ids,
            "invalid_id_format": self.invalid_id_format,
            "empty_core_rows": self.empty_core_rows,
            "sheet_name": self.sheet_name,
            "columns_used": self.columns_used,
            "columns_dropped": self.columns_dropped,
            "warnings": self.warnings,
        }


@dataclass
class ParseResult:
    cases: list[CaseRecord]
    report: ValidationReport


# --- Text cleaning ---------------------------------------------------------------------

_X000D = "_x000D_"  # literal escape openpyxl surfaces for carriage returns


def clean_cell_text(value: object) -> str:
    """Normalize a single cell value to clean text. Returns "" for empty/None."""
    if value is None:
        return ""
    text = value if isinstance(value, str) else str(value)

    # Strip the literal _x000D_ escape (often "_x000D_\n" for a CRLF -> keep the \n).
    text = text.replace(_X000D, "")
    # Normalize line endings.
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    # Tabs -> single space (preserve text, drop layout tabs).
    text = text.replace("\t", " ")
    # Normalize common bullet glyphs to "- " at line starts.
    lines = []
    for line in text.split("\n"):
        stripped = line.strip()
        stripped = re.sub(r"^[·•▪◦‣]\s*", "- ", stripped)
        # Collapse runs of internal spaces (not newlines) to single spaces.
        stripped = re.sub(r"[ ]{2,}", " ", stripped)
        lines.append(stripped)
    text = "\n".join(lines)
    # Collapse 3+ blank lines to a single blank line; trim outer whitespace.
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    return text


def _normalize_header(value: object) -> str:
    return clean_cell_text(value).strip()


def _is_dropped(header_norm: str) -> bool:
    cf = header_norm.casefold()
    if cf in DROP_COLUMNS:
        return True
    return any(cf.startswith(p) for p in DROP_PREFIXES)


# --- Workbook handling -----------------------------------------------------------------


def _open_workbook(source: bytes | str | Path) -> openpyxl.Workbook:
    try:
        if isinstance(source, (bytes, bytearray)):
            stream = io.BytesIO(bytes(source))
            return openpyxl.load_workbook(stream, read_only=True, data_only=True)
        return openpyxl.load_workbook(filename=str(source), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises a variety of errors on bad input
        raise ExcelParseError(f"Could not open workbook: {exc}") from exc


def _select_sheet(wb: openpyxl.Workbook):
    """Pick the data sheet: preferred name, else the first sheet with an ID header."""
    names = wb.sheetnames
    # Preferred name (normalized match).
    for name in names:
        if name.strip().casefold() == PREFERRED_SHEET.casefold():
            return wb[name]
    # Fallback: first sheet whose header row contains the key column.
    for name in names:
        ws = wb[name]
        header_row, _ = _find_header_row(ws)
        if header_row is not None:
            return ws
    raise ExcelParseError(
        f"No sheet with an '{KEY_COLUMN}' header column found (sheets: {names})"
    )


def _find_header_row(ws, max_scan: int = 10) -> tuple[int | None, list[str] | None]:
    """Locate the header row: the first row (within max_scan) containing the key column.

    Returns (1-based row index, list of normalized headers) or (None, None).
    """
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        headers = [_normalize_header(c) for c in row]
        if any(h.casefold() == KEY_COLUMN.casefold() for h in headers):
            return idx, headers
    return None, None


# --- Public API ------------------------------------------------------------------------


def parse_excel_cases(source: bytes | str | Path) -> ParseResult:
    """Parse a clinical-case workbook into per-row :class:`CaseRecord` objects.

    Args:
        source: raw .xlsx bytes, or a path to the file.

    Returns:
        ParseResult(cases, report). ``cases`` excludes fully-blank rows.

    Raises:
        ExcelParseError: workbook unreadable, or no usable sheet / ID column.
    """
    wb = _open_workbook(source)
    try:
        ws = _select_sheet(wb)
        header_row_idx, headers = _find_header_row(ws)
        if header_row_idx is None or headers is None:
            raise ExcelParseError(f"No '{KEY_COLUMN}' header column on sheet '{ws.title}'")

        report = ValidationReport(sheet_name=ws.title)

        # Resolve column roles by index.
        key_idx: int | None = None
        content_cols: list[tuple[int, str]] = []  # (col_index, label)
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            if header.casefold() == KEY_COLUMN.casefold():
                key_idx = col_idx
                continue
            if _is_dropped(header):
                report.columns_dropped.append(header)
                continue
            content_cols.append((col_idx, header))

        if key_idx is None:
            raise ExcelParseError(f"'{KEY_COLUMN}' column not resolved on sheet '{ws.title}'")

        report.columns_used = [label for _, label in content_cols]

        cases: list[CaseRecord] = []
        seen_ids: dict[str, int] = {}

        data_row_no = 0
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            # Skip entirely empty rows (trailing/blank).
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            data_row_no += 1
            report.total_data_rows += 1

            warnings: list[str] = []

            # --- ID / label ---
            raw_id_cell = row[key_idx] if key_idx < len(row) else None
            raw_id = clean_cell_text(raw_id_cell)
            if not raw_id:
                report.missing_ids.append(data_row_no)
                label = f"Row {data_row_no}"
                warnings.append("missing ID; using row-number label")
            else:
                label = raw_id
                if not ID_PATTERN.match(raw_id):
                    report.invalid_id_format.append(raw_id)
                    warnings.append("ID does not match OLC-CHI-######### format")
                if raw_id in seen_ids:
                    report.duplicate_ids.append(raw_id)
                    warnings.append(f"duplicate ID (first seen at data row {seen_ids[raw_id]})")
                    # Disambiguate label so cases remain distinct.
                    label = f"{raw_id} (dup row {data_row_no})"
                else:
                    seen_ids[raw_id] = data_row_no

            # --- labeled-section concatenation ---
            sections: list[str] = []
            has_core = False
            for col_idx, col_label in content_cols:
                cell = row[col_idx] if col_idx < len(row) else None
                cleaned = clean_cell_text(cell)
                if not cleaned:
                    continue  # per-row empty-skip: no dangling "Label:" sections
                sections.append(f"{col_label}:\n{cleaned}")
                if col_label.casefold() in CORE_COLUMNS:
                    has_core = True

            if not sections:
                # Row had an ID (or content elsewhere) but no usable content columns.
                report.skipped_blank_rows += 1
                continue

            if not has_core:
                report.empty_core_rows.append(label)
                warnings.append("no core clinical text (Diagnosen/Operationen/Verlauf/Procedere)")

            input_text = "\n\n".join(sections)
            cases.append(
                CaseRecord(
                    case_label=label,
                    input_text=input_text,
                    row_number=data_row_no,
                    raw_id=raw_id or None,
                    warnings=warnings,
                )
            )

        report.valid_cases = len(cases)
        if not cases:
            report.warnings.append("no non-empty data rows produced a case")
        return ParseResult(cases=cases, report=report)
    finally:
        wb.close()
